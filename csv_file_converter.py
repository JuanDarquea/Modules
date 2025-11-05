import pandas as pd
import os
from tkinter import filedialog
from tkinter import Tk
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font, Color

def format_excel_file(file_path):
    """Format excel file with borders, alignment and adjusted column width
    Special formatting for header row: black background, white text, centered, uppercase
    
    Args:
    file_path(str): Path to excel file
    """
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active

    # Define black border style
    thin_border = Border(
        left=Side(style='thin', color=Color(rgb='00000000')),
        right=Side(style='thin', color=Color(rgb='00000000')),
        top=Side(style='thin', color=Color(rgb='00000000')),
        bottom=Side(style='thin', color=Color(rgb='00000000'))
    )

    # Define header cell alignment(horizontal center, vertical middle)
    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    # Define alignment for data cells(left horizontal, middle vertical, wrap text)
    alignment = Alignment(
        horizontal='left',
        vertical='center',
        wrap_text=True
    )

    # Define data fill for header
    black_fill=PatternFill(
        start_color=Color(rgb='00000000'), # FF = alpha, 000000 = black
        end_color=Color(rgb='00000000'),
        fill_type='solid'
    )

    # Define white font for header(bold)
    white_font = Font(
        color=Color(rgb='00FFFFFF'), # FF = alpha, FFFFFF = white
        bold=True,
        size=11
    )

    # Get dimensions of data
    max_row = ws.max_row
    max_col = ws.max_column

    # Apply formatting to header cells(row 1)
    for col in range(1, max_col +1):
        cell = ws.cell(row=1, column=col)
        # Apply background
        cell.fill = black_fill
        # Apply white font
        cell.font = white_font
        # Apply center alignment
        cell.alignment = header_alignment
        # Apply borders
        cell.border = thin_border
        # Convert text to uppercase
        if cell.value:
            cell.value = str(cell.value).upper()

    # Apply formatting to all cells with data(row 2 onwards)
    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            # Apply borders
            cell.border = thin_border
            # Apply alignment
            cell.alignment = alignment

    # Set column widths based on your requirements
    # Columns 2 and 3: 3x size (45 units)
    # Columns 1 and 4: 2x size (30 units)
    # Other columns: default (15 units)
    for col in range(1, max_col + 1):
        col_letter = ws.cell(row=1, column=col).column_letter

        if col in [2, 3]: # For columns 2 and 3
            ws.column_dimensions[col_letter].width = 45
        elif col in [1, 4]: # For columns 1and 4
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 15

    # Save the formatted workbook
    wb.save(file_path)
    wb.close

def convert_csv_to_xlsx(csv_path):
    """
    Convert CSV file to Excel with the same filename
    
    Args:
        csv_path (str): Full path to the CSV file
    """
    try:
        # Read CSV file
        df = pd.read_csv(csv_path)

        # Get the filename without extension
        file_dir = os.path.dirname(csv_path)
        file_name = os.path.splitext(os.path.basename(csv_path))[0]

        # Create output path with xlsx extension
        output_path = os.path.join(file_dir, f"{file_name}.xlsx")

        # Convert to excel
        df.to_excel(output_path, index=False)

        # Apply formatting
        format_excel_file(output_path)

        print("\nFile successfully converted and formatted")
        print(f"Input file = {csv_path}")
        print(f"Output file = {output_path}")

        print(f"\nFormatting applied:")
        print(f"  Header Row:")
        print(f"    • Black background")
        print(f"    • White, bold text")
        print(f"    • Center aligned (horizontal & vertical)")
        print(f"    • All caps")
        print(f"  Data Rows:")
        print(f"    • Black borders on all cells")
        print(f"    • Left-aligned, vertically centered text")
        print(f"    • Text wrapping enabled")
        print(f"  Column Widths:")
        print(f"    • Columns 1 & 4: 30 units wide")
        print(f"    • Columns 2 & 3: 45 units wide")

        return output_path
    
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exec()
        return None
    
def select_csv_file():
    """Open file dialog to select a CSV file"""
    root = Tk()
    root.withdraw() # Hide the main window
    root.attributes('-topmost', True) # Bring dialog to front

    # Open file dialog
    file_path = filedialog.askopenfilename(
        title="Select CSV file to convert", 
        filetypes=[
            ("CSV files", "*.csv"), 
            ("All files", "*.*")
        ],
        initialdir=os.path.expanduser("~") # Start in users home directory
    )

    root.destroy() # Close hidden window

    return file_path

def main():
    """Main execution function"""
    print("Please select a CSV file to convert to excel.")

    # Get filepath from dialog
    csv_path = select_csv_file()

    # Get the filename without extension
    file_dir = os.path.dirname(csv_path)
    file_name = os.path.splitext(os.path.basename(csv_path))[0]

    if csv_path:
        print(f"\nSelected file: {file_name}",
              f"\nFrom directory: {file_dir}")
        convert_csv_to_xlsx(csv_path)
    else:
        print("No file selected. Exiting!")

if __name__ == "__main__":
    main()